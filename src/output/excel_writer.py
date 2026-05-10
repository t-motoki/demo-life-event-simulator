"""Excel出力モジュール（openpyxl使用）"""

from datetime import date, datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.domain.cashflow_analysis import analyze
from src.domain.models import CashFlowRow, HousingEvent, IncomeModel, Scenario


# 貯蓄残高マイナス時の背景色
NEGATIVE_SAVINGS_FILL = PatternFill(
    start_color="FF9999",
    end_color="FF9999",
    fill_type="solid",
)

# マイナス金額の文字色
NEGATIVE_FONT_COLOR = "FF0000"  # 赤



def _write_cashflow_sheet(
    ws,
    rows: list[CashFlowRow],
    has_spouse: bool,
) -> None:
    """キャッシュフロー一覧シートに書き込む"""

    # --- ヘッダー定義 ---
    headers = ["年（西暦）", "年齢（本人）"]
    if has_spouse:
        headers.append("年齢（配偶者）")
    headers += [
        "収入合計",
        "支出合計",
        "住宅ローン控除",
        "年間収支",
        "貯蓄残高",
        "イベント",
    ]

    # ヘッダー行書き込み
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # --- データ行書き込み ---
    for row_idx, row in enumerate(rows, start=2):
        col = 1
        ws.cell(row=row_idx, column=col, value=row.year)
        col += 1
        ws.cell(row=row_idx, column=col, value=row.age_client)
        col += 1

        if has_spouse:
            ws.cell(row=row_idx, column=col, value=row.age_spouse)
            col += 1

        # 金額列
        amount_cols: list[tuple[int, int]] = []  # (col_idx, value)
        for value in [
            row.income_total,
            row.expense_total,
            row.loan_deduction,
            row.net,
            row.savings,
        ]:
            amount_cols.append((col, value))
            col += 1

        # イベントラベル
        ws.cell(row=row_idx, column=col, value=row.events_label)

        # 金額セルのフォーマット・色設定
        for col_idx, value in amount_cols:
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.number_format = '#,##0'
            cell.alignment = Alignment(horizontal="right")
            if value < 0:
                cell.font = Font(color=NEGATIVE_FONT_COLOR)

        # 貯蓄残高がマイナスの行は背景色を変更
        if row.savings < 0:
            for c in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=c).fill = NEGATIVE_SAVINGS_FILL

    # --- 列幅の自動調整 ---
    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        # ヘッダー長に基づく最小幅
        ws.column_dimensions[col_letter].width = max(len(header) * 2, 12)

    # イベント列は広めに
    event_col = get_column_letter(len(headers))
    ws.column_dimensions[event_col].width = 30


def _write_input_summary_sheet(ws, scenario: Scenario) -> None:
    """入力内容確認シートに書き込む"""
    ws.title = "入力内容確認"

    rows = [
        ("シミュレーション開始年", scenario.start_year),
        ("シミュレーション終了年齢（本人）", scenario.end_age),
        ("", ""),
        ("【本人】", ""),
        ("現在年齢", scenario.client.age),
        ("税引後年収", scenario.client.annual_income),
        ("収入モデル", scenario.client.income_model.value),
        ("昇給率", scenario.client.raise_rate),
        ("定年年齢", scenario.client.retirement_age),
        ("定年後年収", scenario.client.post_retirement_income),
        ("年金開始年齢", scenario.client.pension_start_age),
        ("年間年金額", scenario.client.pension_annual),
        ("", ""),
    ]

    if scenario.spouse is not None:
        rows += [
            ("【配偶者】", ""),
            ("現在年齢", scenario.spouse.age),
            ("税引後年収", scenario.spouse.annual_income),
            ("収入モデル", scenario.spouse.income_model.value),
            ("定年年齢", scenario.spouse.retirement_age),
            ("定年後年収", scenario.spouse.post_retirement_income),
            ("年金開始年齢", scenario.spouse.pension_start_age),
            ("年間年金額", scenario.spouse.pension_annual),
            ("", ""),
        ]

    for event in scenario.events:
        if isinstance(event, HousingEvent):
            age_at_purchase = scenario.client.age + (event.year - scenario.start_year)
            payoff_age = age_at_purchase + event.loan_years
            rows += [
                ("【住宅ローン】", ""),
                ("物件価格", event.price),
                ("頭金", event.down_payment),
                ("借入年数", event.loan_years),
                ("金利", event.interest_rate),
                ("住宅ローン控除", "あり" if event.use_tax_deduction else "なし"),
                ("完済年齢", payoff_age),
                ("", ""),
            ]

    rows += [
        ("【月間固定費】", ""),
        ("生活費", scenario.monthly_expenses.living),
        ("保険料", scenario.monthly_expenses.insurance),
        ("その他", scenario.monthly_expenses.other),
        ("", ""),
        ("【初期貯蓄残高】", scenario.savings_initial),
    ]

    for r_idx, (label, value) in enumerate(rows, start=1):
        ws.cell(row=r_idx, column=1, value=label)
        cell = ws.cell(row=r_idx, column=2, value=value)
        if isinstance(value, int) and label not in ("現在年齢", "定年年齢", "年金開始年齢", "シミュレーション開始年", "シミュレーション終了年齢（本人）", "借入年数", "完済年齢"):
            cell.number_format = '#,##0'
            cell.alignment = Alignment(horizontal="right")

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20


def _write_notes_sheet(
    ws,
    scenario: Scenario,
    rows: list[CashFlowRow],
    fp_comment: str,
    fp_name: str,
) -> None:
    """「前提条件・注釈」シートに書き込む。

    analyze() を内部で呼び出して貯蓄最低値・赤字期間を取得する。
    """
    ws.title = "前提条件・注釈"
    summary = analyze(rows)
    today_str = date.today().isoformat()

    # 収入モデルの説明テキスト
    income_model = scenario.client.income_model
    if income_model == IncomeModel.FLAT:
        income_desc = "収入一定（毎年同額の収入）"
    elif income_model == IncomeModel.RAISE_RATE:
        rate_pct = scenario.client.raise_rate * 100
        income_desc = f"昇給率 {rate_pct:.1f}%（毎年昇給）"
    else:  # POST_RETIREMENT
        income_desc = f"定年後減額（定年後年収: {scenario.client.post_retirement_income:,}円）"

    # --- シート構成 ---
    # 1行目: 作成日
    ws.cell(row=1, column=1, value="作成日").font = Font(bold=True)
    ws.cell(row=1, column=2, value=today_str)

    # 2行目: FP 名
    ws.cell(row=2, column=1, value="FP名").font = Font(bold=True)
    ws.cell(row=2, column=2, value=fp_name if fp_name else "")

    # 3行目: 空白
    # （何も書かない）

    # 4行目: 【前提条件】見出し
    ws.cell(row=4, column=1, value="【前提条件】").font = Font(bold=True)

    # 5行目: 収入モデル
    ws.cell(row=5, column=1, value="収入モデル").font = Font(bold=True)
    ws.cell(row=5, column=2, value=income_desc)

    # 6行目: インフレ率
    ws.cell(row=6, column=1, value="インフレ率").font = Font(bold=True)
    ws.cell(row=6, column=2, value="0%（固定）")

    # 7行目: 運用利回り
    ws.cell(row=7, column=1, value="運用利回り").font = Font(bold=True)
    ws.cell(row=7, column=2, value="0%（固定）")

    # 8行目: 住宅ローン金利（HousingEvent がある場合のみ）
    current_row = 8
    for event in scenario.events:
        if isinstance(event, HousingEvent):
            rate_pct = event.interest_rate * 100
            ws.cell(row=current_row, column=1, value="住宅ローン金利").font = Font(bold=True)
            ws.cell(row=current_row, column=2, value=f"{rate_pct:.1f}%")
            current_row += 1
            break

    # 空白行
    current_row += 1

    # 【注目ポイント】見出し
    ws.cell(row=current_row, column=1, value="【注目ポイント】").font = Font(bold=True)
    current_row += 1

    # 貯蓄最低値
    savings_low_text = (
        f"{summary.savings_low.year}年"
        f"（{summary.savings_low.age}歳）・"
        f"{summary.savings_low.amount:,}円"
    )
    ws.cell(row=current_row, column=1, value="貯蓄残高最低時期").font = Font(bold=True)
    ws.cell(row=current_row, column=2, value=savings_low_text)
    current_row += 1

    # 赤字期間（あれば各行に書く）
    if summary.deficit_periods:
        for period in summary.deficit_periods:
            if period.start_year == period.end_year:
                period_text = f"{period.start_year}年"
            else:
                period_text = f"{period.start_year}〜{period.end_year}年"
            ws.cell(row=current_row, column=1, value="赤字期間").font = Font(bold=True)
            ws.cell(row=current_row, column=2, value=period_text)
            current_row += 1

    # 空白行
    current_row += 1

    # FP コメント見出し
    ws.cell(row=current_row, column=1, value="FPコメント").font = Font(bold=True)
    current_row += 1

    # FP コメント本文（Klee One フォント）
    comment_cell = ws.cell(row=current_row, column=2, value=fp_comment if fp_comment else "")
    comment_cell.font = Font(name="Klee One")

    # 列幅の調整
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 50


def write_excel(
    scenario: Scenario,
    rows: list[CashFlowRow],
    output_dir: str | Path = ".",
    fp_comment: str = "",
    fp_name: str = "",
) -> Path:
    """キャッシュフロー結果をExcelファイルに書き出す

    Args:
        scenario: シナリオ情報
        rows: キャッシュフロー一覧
        output_dir: 出力先ディレクトリ
        fp_comment: FP コメント文字列（デフォルト空文字・後方互換）
        fp_name: FP 名（デフォルト空文字・後方互換）

    Returns:
        生成されたExcelファイルのパス
    """
    wb = Workbook()

    # シート1: キャッシュフロー一覧
    ws_cf = wb.active
    ws_cf.title = "キャッシュフロー一覧"
    has_spouse = scenario.spouse is not None
    _write_cashflow_sheet(ws_cf, rows, has_spouse)

    # シート2: 入力内容確認
    ws_input = wb.create_sheet("入力内容確認")
    _write_input_summary_sheet(ws_input, scenario)

    # シート3: 前提条件・注釈
    ws_notes = wb.create_sheet("前提条件・注釈")
    _write_notes_sheet(ws_notes, scenario, rows, fp_comment, fp_name)

    # ファイル名: cf_simulation_YYYYMMDD.xlsx
    today = datetime.now().strftime("%Y%m%d")
    filename = f"cf_simulation_{today}.xlsx"
    output_path = Path(output_dir) / filename

    wb.save(output_path)
    return output_path
