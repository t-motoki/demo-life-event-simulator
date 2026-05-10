"""Excel出力のテスト"""

import openpyxl
import pytest

from src.domain.cashflow import simulate
from src.domain.models import (
    CashFlowRow,
    Client,
    HousingEvent,
    IncomeModel,
    MonthlyExpenses,
    Scenario,
)
from src.output.excel_writer import write_excel


def _base_scenario(**kwargs) -> Scenario:
    defaults = dict(
        client=Client(
            age=35,
            annual_income=5_000_000,
            income_model=IncomeModel.FLAT,
            retirement_age=65,
            post_retirement_income=0,
            pension_start_age=65,
            pension_annual=0,
        ),
        spouse=None,
        savings_initial=3_000_000,
        end_age=70,
        monthly_expenses=MonthlyExpenses(living=200_000, insurance=20_000, other=10_000),
        events=[],
        start_year=2025,
    )
    defaults.update(kwargs)
    return Scenario(**defaults)


def _read_input_sheet(tmp_path, scenario):
    """入力内容確認シートの内容を {ラベル: 値} で返す"""
    rows = simulate(scenario)
    path = write_excel(scenario, rows, output_dir=tmp_path)
    wb = openpyxl.load_workbook(path)
    ws = wb["入力内容確認"]
    return {row[0].value: row[1].value for row in ws.iter_rows() if row[0].value}


class TestInputSummaryHousingLoan:
    def test_payoff_age_present_in_sheet(self, tmp_path):
        """住宅ローンがある場合、完済年齢が入力確認シートに出力される"""
        housing = HousingEvent(
            year=2027,  # 借入時 client 年齢: 35 + (2027-2025) = 37歳
            price=40_000_000,
            down_payment=8_000_000,
            loan_years=25,  # 完済年齢: 37 + 25 = 62歳
            interest_rate=0.015,
            use_tax_deduction=True,
        )
        scenario = _base_scenario(events=[housing])
        data = _read_input_sheet(tmp_path, scenario)
        assert "完済年齢" in data
        assert data["完済年齢"] == 62

    def test_payoff_age_calculation(self, tmp_path):
        """完済年齢 = 借入時年齢 + loan_years で計算される"""
        # client 35歳 / start_year=2025 / housing year=2025 → 借入時35歳
        # loan_years=20 → 完済年齢=55
        housing = HousingEvent(
            year=2025,
            price=30_000_000,
            down_payment=5_000_000,
            loan_years=20,
            interest_rate=0.015,
            use_tax_deduction=True,
        )
        scenario = _base_scenario(events=[housing])
        data = _read_input_sheet(tmp_path, scenario)
        assert data["完済年齢"] == 55

    def test_no_housing_event_no_payoff_age(self, tmp_path):
        """住宅イベントがない場合、完済年齢は出力されない"""
        scenario = _base_scenario()
        data = _read_input_sheet(tmp_path, scenario)
        assert "完済年齢" not in data


# ---------------------------------------------------------------------------
# 前提条件・注釈シート（「前提条件・注釈」第3シート）のテスト（AC-1-x）
# ---------------------------------------------------------------------------

def _make_cf_row(year: int, age_client: int, net: int, savings: int) -> CashFlowRow:
    """テスト用 CashFlowRow を簡易生成する"""
    return CashFlowRow(
        year=year,
        age_client=age_client,
        age_spouse=None,
        income_total=5_000_000,
        expense_total=4_000_000,
        loan_deduction=0,
        net=net,
        savings=savings,
        events_label="",
    )


def _get_notes_sheet_text(tmp_path, scenario, rows=None, fp_comment="", fp_name="") -> str:
    """前提条件・注釈シートのテキスト全体を結合して返すヘルパー"""
    if rows is None:
        rows = simulate(scenario)
    path = write_excel(scenario, rows, output_dir=tmp_path, fp_comment=fp_comment, fp_name=fp_name)
    wb = openpyxl.load_workbook(path)
    ws = wb["前提条件・注釈"]
    texts = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value:
                texts.append(str(cell.value))
    return "\n".join(texts)


def _get_notes_sheet_cells(tmp_path, scenario, rows=None, fp_comment="", fp_name=""):
    """前提条件・注釈シートの openpyxl worksheet を返すヘルパー"""
    if rows is None:
        rows = simulate(scenario)
    path = write_excel(scenario, rows, output_dir=tmp_path, fp_comment=fp_comment, fp_name=fp_name)
    wb = openpyxl.load_workbook(path)
    return wb["前提条件・注釈"]


class TestNotesSheet:
    def test_notes_sheet_exists_as_third_sheet(self, tmp_path):
        """「前提条件・注釈」シートが3枚目として追加されている（AC-1-1）"""
        scenario = _base_scenario()
        rows = simulate(scenario)
        path = write_excel(scenario, rows, output_dir=tmp_path)
        wb = openpyxl.load_workbook(path)

        assert len(wb.sheetnames) == 3
        assert wb.sheetnames[0] == "キャッシュフロー一覧"
        assert wb.sheetnames[1] == "入力内容確認"
        assert wb.sheetnames[2] == "前提条件・注釈"

    def test_notes_sheet_creation_date(self, tmp_path):
        """作成日が実行日の日付になっている（AC-1-2）"""
        from datetime import date

        scenario = _base_scenario()
        text = _get_notes_sheet_text(tmp_path, scenario)
        today = date.today()

        # ISO 形式または日本語形式どちらかが含まれている
        iso_format = today.isoformat()  # 2026-05-10
        ja_format = f"{today.year}年{today.month}月{today.day}日"
        assert iso_format in text or ja_format in text

    def test_notes_sheet_fp_name_default_empty(self, tmp_path):
        """FP 名を指定しない場合、FP 名セルが空文字列または空白のみ（AC-1-3）"""
        scenario = _base_scenario()
        ws = _get_notes_sheet_cells(tmp_path, scenario)

        # "FP名" ラベルの隣セルを確認
        fp_name_value = None
        for row in ws.iter_rows():
            for i, cell in enumerate(row):
                if cell.value and "FP名" in str(cell.value):
                    # 隣のセル（B列）を確認
                    next_cell = row[i + 1] if i + 1 < len(row) else None
                    if next_cell:
                        fp_name_value = next_cell.value
        # 空文字列 or None（空白のみ）
        assert fp_name_value is None or str(fp_name_value).strip() == ""

    def test_notes_sheet_income_model_flat_text(self, tmp_path):
        """収入モデル flat の説明文が含まれる（AC-1-4）"""
        scenario = _base_scenario(client=Client(
            age=35,
            annual_income=5_000_000,
            income_model=IncomeModel.FLAT,
        ))
        text = _get_notes_sheet_text(tmp_path, scenario)
        assert "収入一定" in text or "毎年同額" in text

    def test_notes_sheet_income_model_raise_rate_text(self, tmp_path):
        """収入モデル raise_rate の説明文に昇給率が含まれる（AC-1-5）"""
        scenario = _base_scenario(client=Client(
            age=35,
            annual_income=5_000_000,
            income_model=IncomeModel.RAISE_RATE,
            raise_rate=0.02,
        ))
        text = _get_notes_sheet_text(tmp_path, scenario)
        assert "昇給率" in text
        assert "2.0%" in text or "2%" in text

    def test_notes_sheet_income_model_post_retirement_text(self, tmp_path):
        """収入モデル post_retirement の説明文に定年後・減額が含まれる（AC-1-6）"""
        scenario = _base_scenario(client=Client(
            age=35,
            annual_income=5_000_000,
            income_model=IncomeModel.POST_RETIREMENT,
            retirement_age=60,
            post_retirement_income=1_500_000,
        ))
        text = _get_notes_sheet_text(tmp_path, scenario)
        assert "定年後" in text or "減額" in text

    def test_notes_sheet_inflation_assumption_text(self, tmp_path):
        """インフレ率 0%・運用利回り 0% の固定前提テキストが含まれる（AC-1-7）"""
        scenario = _base_scenario()
        text = _get_notes_sheet_text(tmp_path, scenario)
        assert "インフレ率" in text or "インフレ" in text
        assert "運用利回り" in text or "利回り" in text

    def test_notes_sheet_housing_interest_rate_with_event(self, tmp_path):
        """HousingEvent あり時に住宅ローン金利の記載がある（AC-1-8）"""
        housing = HousingEvent(
            year=2027,
            price=40_000_000,
            down_payment=8_000_000,
            loan_years=25,
            interest_rate=0.015,
            use_tax_deduction=True,
        )
        scenario = _base_scenario(events=[housing])
        text = _get_notes_sheet_text(tmp_path, scenario)
        assert "住宅ローン金利" in text or "ローン金利" in text
        assert "1.5%" in text

    def test_notes_sheet_no_housing_interest_rate_text(self, tmp_path):
        """HousingEvent なし時は住宅ローン金利の記載がない（AC-1-9）"""
        scenario = _base_scenario()
        text = _get_notes_sheet_text(tmp_path, scenario)
        assert "住宅ローン金利" not in text

    def test_notes_sheet_savings_low_values(self, tmp_path):
        """貯蓄最低値の年・年齢・金額が注目ポイント欄に含まれる（AC-1-10）"""
        # 2028年・37歳が最低貯蓄（ net がマイナス最大の行）
        rows = [
            _make_cf_row(2026, 35, 100_000, 3_100_000),
            _make_cf_row(2027, 36, 50_000, 3_150_000),
            _make_cf_row(2028, 37, -2_000_000, 1_150_000),  # 最低
            _make_cf_row(2029, 38, 200_000, 1_350_000),
            _make_cf_row(2030, 39, 300_000, 1_650_000),
        ]
        scenario = _base_scenario()
        text = _get_notes_sheet_text(tmp_path, scenario, rows=rows)
        assert "2028" in text
        assert "37" in text
        # 金額（カンマ区切り含む）
        assert "1,150,000" in text or "115万" in text or "1150000" in text

    def test_notes_sheet_savings_low_monotonic(self, tmp_path):
        """単調増加シナリオでは最初の行が savings_low（AC-1-11）"""
        rows = [
            _make_cf_row(2026, 35, 100_000, 3_100_000),
            _make_cf_row(2027, 36, 200_000, 3_300_000),
            _make_cf_row(2028, 37, 300_000, 3_600_000),
        ]
        scenario = _base_scenario()
        # エラーが発生しないことを確認
        text = _get_notes_sheet_text(tmp_path, scenario, rows=rows)
        assert "2026" in text  # 最初の年が savings_low

    def test_notes_sheet_deficit_period_listed(self, tmp_path):
        """赤字期間が注目ポイント欄に明示されている（AC-1-12）"""
        rows = [
            _make_cf_row(2045, 54, 100_000, 3_000_000),
            _make_cf_row(2046, 55, -100_000, 2_900_000),
            _make_cf_row(2047, 56, -200_000, 2_700_000),
            _make_cf_row(2048, 57, -150_000, 2_550_000),
            _make_cf_row(2049, 58, 100_000, 2_650_000),
        ]
        scenario = _base_scenario()
        text = _get_notes_sheet_text(tmp_path, scenario, rows=rows)
        assert "2046" in text or "赤字" in text

    def test_notes_sheet_no_deficit_text_when_all_positive(self, tmp_path):
        """全年 net >= 0 のとき赤字期間に関する記載がないか「赤字期間なし」（AC-1-13）"""
        rows = [
            _make_cf_row(2026, 35, 100_000, 3_100_000),
            _make_cf_row(2027, 36, 200_000, 3_300_000),
        ]
        scenario = _base_scenario()
        text = _get_notes_sheet_text(tmp_path, scenario, rows=rows)
        # 「赤字」という文字がないか、「赤字期間なし」が含まれる
        assert "赤字" not in text or "赤字期間なし" in text

    def test_notes_sheet_fp_comment_font_is_klee_one(self, tmp_path):
        """FP コメントセルのフォントが Klee One（AC-1-14）"""
        scenario = _base_scenario()
        rows = simulate(scenario)
        path = write_excel(
            scenario, rows, output_dir=tmp_path, fp_comment="テストFPコメント", fp_name="山田太郎"
        )
        wb = openpyxl.load_workbook(path)
        ws = wb["前提条件・注釈"]

        # FP コメント本文セルのフォント名を確認
        comment_cell_font_name = None
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == "テストFPコメント":
                    comment_cell_font_name = cell.font.name
        assert comment_cell_font_name == "Klee One"

    def test_notes_sheet_fp_comment_empty_is_blank(self, tmp_path):
        """FP コメントが空のとき欄は空のままでエラーが出ない（AC-1-15）"""
        scenario = _base_scenario()
        rows = simulate(scenario)
        # 空文字列を渡してもエラーが発生しないことを確認
        path = write_excel(scenario, rows, output_dir=tmp_path, fp_comment="", fp_name="")
        wb = openpyxl.load_workbook(path)
        ws = wb["前提条件・注釈"]
        # シートが存在することのみ確認（空コメントでも正常に書き込まれる）
        assert ws is not None
